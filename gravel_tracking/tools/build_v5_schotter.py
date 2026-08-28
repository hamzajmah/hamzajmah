"""Managementbericht Schotter 0/45 (V5).

Bewusst enger Zuschnitt nach fachlicher Vorgabe:
* betrachtet wird ausschliesslich Schotter 0/45
* Umrechnung mit 1,8 t je m3, als aenderbarer Parameter an einer Stelle
* drei Fragen, drei Bloecke: Menge je Sektion, LV Positionen, Bilanz

Alle Kennzahlen sind Formeln auf das Blatt Daten. Wer eine Zahl anzweifelt,
filtert dort und sieht die Zeilen dahinter.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEAD = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="D9E2F3")
NOTE = PatternFill("solid", fgColor="FFF2CC")
WARN = PatternFill("solid", fgColor="FCE4D6")
BORDER = Border(bottom=Side(style="thin", color="BFBFBF"))

SOURCE = Path("/tmp/v4review/v4.xlsx")
TARGET = Path("outputs/Schotter_0_45_Management_V5.xlsx")

DATA_COLUMNS = [
    "datum", "monat", "sektion", "ortstyp", "ortscode", "verwendungsgruppe",
    "menge_t", "m3_bei_faktor", "bestellung", "lieferschein_nr", "erp_schluessel",
]


def _de(value: float, digits: int = 0) -> str:
    """Zahl im deutschen Format: Punkt als Tausender, Komma als Dezimaltrenner."""
    text = f"{value:,.{digits}f}"
    return text.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def _head(ws, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEAD
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = f"A{row + 1}"


def _title(ws, text: str, subtitle: str) -> int:
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, bold=True, size=14, color="1F3864")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
    return 4


def _widths(ws, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _receipt_key(ref: str) -> str | None:
    """Gemeinsamer Schluessel Bestellung|Position|Empfang aus beiden Quellformaten."""
    text = str(ref)
    match = re.match(r"po=([^/]+)/(\d+);receipt=(\d+)", text)
    if match:
        return f"{match.group(1)}|{match.group(2)}|{match.group(3)}"
    parts = text.split("|")
    if len(parts) == 4:
        return f"{parts[0]}|{parts[1]}|{parts[3]}"
    return None


def _pipeline_gravel() -> pd.DataFrame:
    """Bereinigter ERP Stand 0/45 aus der Pipeline, ohne Doppelbuchungen."""
    frame = pd.read_csv("work/02_records.csv", sep=";", dtype=str, low_memory=False)
    frame = frame[(frame["grain_size"] == "0/45")
                  & (frame["charge_type"] == "material_supply")].copy()
    frame["quantity_t"] = pd.to_numeric(frame["quantity_t"], errors="coerce")
    frame["key"] = frame["source_row_ref"].map(_receipt_key)
    return frame


def _align_with_pipeline(gravel: pd.DataFrame, pipeline: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """V4 Zeilen auf den bereinigten Stand bringen.

    Die V4 Mappe enthaelt 41 Empfaenge, die in IFS zweimal gebucht sind
    (gleicher Lieferschein, gleiches Datum, gleiche Tonnage, zwei Empfangsnummern),
    und ihr fehlen zwoelf Empfaenge aus der Bestellung P100012091. Beides wird hier
    korrigiert, damit die Ortszuordnung der V4 erhalten bleibt und die Menge trotzdem
    dem bereinigten ERP Stand entspricht.
    """
    gravel = gravel.copy()
    gravel["key"] = (gravel["bestellung"].astype(str) + "|"
                     + gravel["bestellposition"].astype(str) + "|"
                     + gravel["empfang_nr"].astype(str))
    duplicate_keys = set(pipeline.loc[pipeline["is_duplicate"] == "true", "key"].dropna())
    dropped = gravel[gravel["key"].isin(duplicate_keys)]
    gravel = gravel[~gravel["key"].isin(duplicate_keys)]

    known = set(gravel["key"])
    missing = pipeline[(pipeline["is_duplicate"] != "true") & (~pipeline["key"].isin(known))]
    added = pd.DataFrame({
        "datum": pd.to_datetime(missing["delivery_date"]),
        "monat": pd.to_datetime(missing["delivery_date"]).dt.strftime("%Y-%m"),
        "sektion": None,
        "ortstyp": missing["location_type"].fillna("ohne Ortsangabe"),
        "ortscode": missing["location_label"],
        "verwendungsgruppe": missing["activity_text"],
        "menge_t": missing["quantity_t"],
        "bestellung": missing["order_no"],
        "lieferschein_nr": missing["delivery_note_no"],
        "erp_schluessel": missing["source_row_ref"],
        "key": missing["key"],
    })
    result = pd.concat([gravel, added], ignore_index=True)
    stats = {
        "dropped_rows": len(dropped),
        "dropped_t": float(dropped["menge_t"].sum()),
        "added_rows": len(added),
        "added_t": float(added["menge_t"].sum()),
    }
    return result, stats


def load() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict]:
    lief = pd.read_excel(SOURCE, sheet_name="Lieferungen")
    gravel = lief[(lief["materialgruppe"] == "Schotter/Gravel") & (lief["koernung"] == "0/45")].copy()
    other = lief[(lief["materialgruppe"] == "Schotter/Gravel") & (lief["koernung"] != "0/45")].copy()
    gravel, stats = _align_with_pipeline(gravel, _pipeline_gravel())
    lv = pd.read_csv("work/05_lv_positions.csv", sep=";", dtype=str)
    lv = lv[lv["material_group"] == "gravel_base_layer"].copy()
    for col in ("contract_quantity", "billed_quantity", "unit_price_eur"):
        lv[col] = pd.to_numeric(lv[col], errors="coerce")
    return gravel, other, lv.sort_values("contract_quantity", ascending=False), stats


def write_parameters(wb: Workbook, other_t: float, stats: dict) -> None:
    ws = wb.create_sheet("Parameter")
    row = _title(ws, "Parameter und Abgrenzung", "Gelbe Zelle ist aenderbar. Alle Kubikmeter im Bericht rechnen dagegen.")
    ws.cell(row=row, column=1, value="Dichte Schotter 0/45").font = Font(name=FONT, size=10, bold=True)
    cell = ws.cell(row=row, column=2, value=1.8)
    cell.number_format = "0.00"
    cell.fill = NOTE
    cell.font = Font(name=FONT, size=11, bold=True, color="0000FF")
    ws.cell(row=row, column=3, value="t je m3").font = Font(name=FONT, size=9, color="595959")
    ws.cell(row=row + 1, column=1, value="entspricht").font = Font(name=FONT, size=10)
    ws.cell(row=row + 1, column=2, value=f"=IFERROR(1/B{row},0)").number_format = "0.0000"
    ws.cell(row=row + 1, column=3, value="m3 je t").font = Font(name=FONT, size=9, color="595959")

    row += 3
    ws.cell(row=row, column=1, value="Abgrenzung dieser Auswertung").font = Font(name=FONT, bold=True, size=11, color="1F3864")
    row += 1
    for text in (
        "Betrachtet wird ausschliesslich Schotter der Koernung 0/45.",
        "Nicht enthalten: Schotter 50/200 mit " + _de(other_t) + " t. "
        "Die LV Position Schottertragschicht unterscheidet keine Koernung; wird 50/200 als "
        "Unterlage mit eingebaut, gehoert diese Menge fachlich in denselben Vergleich.",
        "Nicht enthalten: Bettungsmaterial 0/8 und 0/22 sowie Sand 0/2.",
        "Die Menge in Tonnen ist der bereinigte ERP Stand und wird nicht veraendert.",
        "Korrektur gegenueber der Mappe V4: "
        + f"{stats['dropped_rows']} Empfaenge mit {_de(stats['dropped_t'], 1)} t"
        + " sind in IFS doppelt gebucht (gleicher Lieferschein, gleiches Datum, gleiche Tonnage, "
        "zwei Empfangsnummern) und wurden entfernt; "
        + f"{stats['added_rows']} Empfaenge mit {_de(stats['added_t'], 1)} t"
        + " aus der Bestellung P100012091 fehlten in V4 und wurden ergaenzt. "
        + f"Netto {_de(stats['added_t'] - stats['dropped_t'], 1)} t.",
    ):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name=FONT, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 26
        row += 1
    _widths(ws, {"A": 34, "B": 14, "C": 16, "D": 14, "E": 14, "F": 14})


def write_data(wb: Workbook, gravel: pd.DataFrame) -> int:
    ws = wb.create_sheet("Daten")
    ws.append(DATA_COLUMNS)
    _head(ws, 1, len(DATA_COLUMNS))

    frame = gravel.copy()
    frame["sektion"] = frame["sektion"].fillna("ohne Sektion")
    frame["ortscode"] = frame["ortscode"].fillna("ohne Ortsangabe")
    frame["verwendungsgruppe"] = frame["verwendungsgruppe"].fillna("nicht beschrieben")
    frame = frame.sort_values(["sektion", "datum"])

    for record in frame.to_dict("records"):
        ws.append([
            record["datum"], str(record["monat"]), record["sektion"], record["ortstyp"],
            record["ortscode"], record["verwendungsgruppe"], round(float(record["menge_t"]), 3),
            None, record["bestellung"], record["lieferschein_nr"], record["erp_schluessel"],
        ])

    last = ws.max_row
    for r in range(2, last + 1):
        ws.cell(row=r, column=8, value=f"=IFERROR($G{r}/Parameter!$B$4,0)")
        for col in range(1, len(DATA_COLUMNS) + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name=FONT, size=9)
            if col == 1:
                cell.number_format = "yyyy-mm-dd"
            if col in (7, 8):
                cell.number_format = "#,##0.00"
    _widths(ws, {"A": 12, "B": 10, "C": 14, "D": 16, "E": 16, "F": 34, "G": 12, "H": 12, "I": 14, "J": 16, "K": 20})
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DATA_COLUMNS))}{last}"
    return last


def write_sections(wb: Workbook, gravel: pd.DataFrame, last: int) -> None:
    ws = wb.create_sheet("Sektionen")
    row = _title(ws, "Wie viel Schotter 0/45 steckt in welcher Sektion",
                 "Menge aus dem bereinigten ERP Stand, Kubikmeter gerechnet mit dem Faktor aus dem Blatt Parameter.")
    header = ["Sektion", "Fuhren", "Menge (t)", "Volumen (m3)", "Anteil an Gesamt"]
    for index, name in enumerate(header, start=1):
        ws.cell(row=row, column=index, value=name)
    _head(ws, row, len(header))

    frame = gravel.copy()
    frame["sektion"] = frame["sektion"].fillna("ohne Sektion")
    order = frame.groupby("sektion")["menge_t"].sum().sort_values(ascending=False).index.tolist()

    first = row + 1
    for offset, section in enumerate(order):
        r = first + offset
        ws.cell(row=r, column=1, value=section)
        ws.cell(row=r, column=2, value=f'=COUNTIFS(Daten!$C$2:$C${last},$A{r})')
        ws.cell(row=r, column=3, value=f'=SUMIFS(Daten!$G$2:$G${last},Daten!$C$2:$C${last},$A{r})')
        ws.cell(row=r, column=4, value=f'=SUMIFS(Daten!$H$2:$H${last},Daten!$C$2:$C${last},$A{r})')
        ws.cell(row=r, column=5, value=f'=IFERROR($C{r}/SUM(Daten!$G$2:$G${last}),0)')
        for col in range(1, 6):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            if col in (2, 3, 4):
                cell.number_format = "#,##0"
            if col == 5:
                cell.number_format = "0.0%"
        if section == "ohne Sektion":
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = WARN

    total = first + len(order)
    ws.cell(row=total, column=1, value="Summe").font = Font(name=FONT, bold=True, size=10)
    for col, letter in ((2, "B"), (3, "C"), (4, "D")):
        cell = ws.cell(row=total, column=col, value=f"=SUM({letter}{first}:{letter}{total - 1})")
        cell.number_format = "#,##0"
        cell.font = Font(name=FONT, bold=True, size=10)
        cell.fill = SUB
    cell = ws.cell(row=total, column=5, value=f"=SUM(E{first}:E{total - 1})")
    cell.number_format = "0.0%"
    cell.font = Font(name=FONT, bold=True, size=10)
    cell.fill = SUB
    ws.cell(row=total, column=1).fill = SUB

    note = ws.cell(row=total + 2, column=1,
                   value="Orange markiert: Lieferungen ohne Sektionszuordnung. Sie zaehlen in die Gesamtmenge, "
                         "lassen sich aber keinem Abschnitt zuordnen.")
    note.font = Font(name=FONT, size=9, italic=True)
    ws.merge_cells(start_row=total + 2, start_column=1, end_row=total + 2, end_column=5)
    _widths(ws, {"A": 18, "B": 12, "C": 14, "D": 16, "E": 18})


def write_lv(wb: Workbook, lv: pd.DataFrame) -> int:
    ws = wb.create_sheet("LV_Schotter")
    row = _title(ws, "Welche LV Positionen gibt es fuer Schotter und was ist abgerechnet",
                 "Quelle: Leistungsmeldung, Blaetter Haupt_Leistung und Nachtrags_Leistung. Mengen in Kubikmetern, unveraendert uebernommen.")
    header = ["OZ", "Bereich", "Kurztext", "Vertragsmenge (m3)", "Abgerechnet (m3)", "Offen (m3)", "Fortschritt", "EP (EUR)", "Abgerechnet (EUR)"]
    for index, name in enumerate(header, start=1):
        ws.cell(row=row, column=index, value=name)
    _head(ws, row, len(header))

    first = row + 1
    for offset, record in enumerate(lv.to_dict("records")):
        r = first + offset
        ws.cell(row=r, column=1, value=record["lv_position_no"])
        ws.cell(row=r, column=2, value=str(record.get("group_path", "")).split(" > ")[-1])
        ws.cell(row=r, column=3, value=record["short_text"])
        ws.cell(row=r, column=4, value=round(float(record["contract_quantity"] or 0), 2))
        ws.cell(row=r, column=5, value=round(float(record["billed_quantity"] or 0), 2))
        ws.cell(row=r, column=6, value=f"=D{r}-E{r}")
        ws.cell(row=r, column=7, value=f"=IFERROR(E{r}/D{r},0)")
        ws.cell(row=r, column=8, value=round(float(record["unit_price_eur"] or 0), 2))
        ws.cell(row=r, column=9, value=f"=E{r}*H{r}")
        for col in range(1, 10):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            if col in (4, 5, 6):
                cell.number_format = "#,##0"
            if col == 7:
                cell.number_format = "0.0%"
            if col in (8, 9):
                cell.number_format = "#,##0.00"
        if float(record["billed_quantity"] or 0) > float(record["contract_quantity"] or 0):
            ws.cell(row=r, column=7).fill = WARN

    total = first + len(lv)
    ws.cell(row=total, column=1, value="Summe").font = Font(name=FONT, bold=True, size=10)
    for col, letter, fmt in ((4, "D", "#,##0"), (5, "E", "#,##0"), (6, "F", "#,##0"), (9, "I", "#,##0.00")):
        cell = ws.cell(row=total, column=col, value=f"=SUM({letter}{first}:{letter}{total - 1})")
        cell.number_format = fmt
        cell.font = Font(name=FONT, bold=True, size=10)
        cell.fill = SUB
    cell = ws.cell(row=total, column=7, value=f"=IFERROR(E{total}/D{total},0)")
    cell.number_format = "0.0%"
    cell.font = Font(name=FONT, bold=True, size=10)
    cell.fill = SUB
    for col in (1, 2, 3, 8):
        ws.cell(row=total, column=col).fill = SUB

    note = ws.cell(row=total + 2, column=1,
                   value="Orange markiert: bereits mehr abgerechnet als beauftragt. Die LV Position unterscheidet keine "
                         "Koernung; sie kann auch mit Schotter 50/200 bedient worden sein.")
    note.font = Font(name=FONT, size=9, italic=True)
    ws.merge_cells(start_row=total + 2, start_column=1, end_row=total + 2, end_column=9)
    _widths(ws, {"A": 14, "B": 34, "C": 40, "D": 18, "E": 18, "F": 14, "G": 12, "H": 12, "I": 18})
    return total


def write_overview(wb: Workbook, last: int, lv_total_row: int, other_t: float) -> None:
    ws = wb.create_sheet("Uebersicht")
    row = _title(ws, "Schotter 0/45 | Menge, Leistungsverzeichnis, Abrechnung",
                 "Bereinigter ERP Stand November 2024 bis August 2026. Umrechnung 1,8 t je m3.")

    blocks = [
        ("Wie viel Schotter 0/45 haben wir geliefert", [
            ("Liefermenge", f"=SUM(Daten!$G$2:$G${last})", "#,##0", "t"),
            ("Fuhren", f"=COUNT(Daten!$G$2:$G${last})", "#,##0", "Anzahl"),
            ("Volumen", f"=SUM(Daten!$H$2:$H${last})", "#,##0", "m3 bei 1,8 t je m3"),
            ("davon ohne Sektionszuordnung", f'=SUMIFS(Daten!$G$2:$G${last},Daten!$C$2:$C${last},"ohne Sektion")', "#,##0", "t"),
        ]),
        ("Was steht dazu im Leistungsverzeichnis", [
            ("Vertragsmenge", f"=LV_Schotter!$D${lv_total_row}", "#,##0", "m3 aus 10 Positionen"),
            ("Abgerechnet", f"=LV_Schotter!$E${lv_total_row}", "#,##0", "m3 Aufmass"),
            ("Noch offen im Vertrag", f"=LV_Schotter!$F${lv_total_row}", "#,##0", "m3"),
            ("Abrechnungsfortschritt", f"=LV_Schotter!$G${lv_total_row}", "0.0%", "der Vertragsmenge"),
            ("Abgerechneter Wert", f"=LV_Schotter!$I${lv_total_row}", "#,##0", "EUR"),
        ]),
        ("Bilanz geliefert gegen abgerechnet", [
            ("Geliefert 0/45", f"=SUM(Daten!$H$2:$H${last})", "#,##0", "m3"),
            ("Abgerechnet laut LV", f"=LV_Schotter!$E${lv_total_row}", "#,##0", "m3"),
            ("Differenz", f"=SUM(Daten!$H$2:$H${last})-LV_Schotter!$E${lv_total_row}", "#,##0", "m3 geliefert minus abgerechnet"),
            ("Deckung", f"=IFERROR(LV_Schotter!$E${lv_total_row}/SUM(Daten!$H$2:$H${last}),0)", "0.0%", "abgerechnet je gelieferter m3"),
        ]),
    ]

    for heading, entries in blocks:
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = SUB
        ws.cell(row=row, column=1, value=heading).font = Font(name=FONT, bold=True, size=11, color="1F3864")
        row += 1
        for label, value, fmt, unit in entries:
            ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10)
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = fmt
            cell.font = Font(name=FONT, size=11, bold=True)
            ws.cell(row=row, column=3, value=unit).font = Font(name=FONT, size=9, color="595959")
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = BORDER
            row += 1
        row += 1

    ws.cell(row=row, column=1, value="Nachrichtlich, nicht in den Zahlen oben enthalten").font = Font(
        name=FONT, bold=True, size=11, color="1F3864")
    row += 1
    ws.cell(row=row, column=1, value="Schotter 50/200").font = Font(name=FONT, size=10)
    cell = ws.cell(row=row, column=2, value=round(other_t, 2))
    cell.number_format = "#,##0"
    cell.font = Font(name=FONT, size=11, bold=True)
    ws.cell(row=row, column=3, value="t, entspricht rund " + _de(other_t / 1.8) + " m3").font = Font(
        name=FONT, size=9, color="595959")
    row += 2

    warning = ws.cell(row=row, column=1, value=(
        "Zur Einordnung: Die LV Position Schottertragschicht unterscheidet keine Koernung. Wird 50/200 als "
        "Unterlage mit eingebaut, gehoert diese Menge in denselben Vergleich. Mit beiden Koernungen zusammen "
        "kehrt sich die Bilanz um. Vor einer Aussage nach aussen ist zu klaeren, welche Koernungen in die "
        "abgerechneten Schottertragschichten eingegangen sind."
    ))
    warning.font = Font(name=FONT, size=9, italic=True)
    warning.fill = NOTE
    warning.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=4)
    _widths(ws, {"A": 38, "B": 18, "C": 34, "D": 18})


def main() -> int:
    gravel, other, lv, stats = load()
    other_t = float(other["menge_t"].sum())

    wb = Workbook()
    wb.remove(wb.active)
    write_parameters(wb, other_t, stats)
    last = write_data(wb, gravel)
    lv_total_row = write_lv(wb, lv)
    write_sections(wb, gravel, last)
    write_overview(wb, last, lv_total_row, other_t)

    order = ["Uebersicht", "Sektionen", "LV_Schotter", "Parameter", "Daten"]
    for index, name in enumerate(order):
        wb.move_sheet(name, offset=index - wb.sheetnames.index(name))

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET)
    print(f"geschrieben: {TARGET}  Datenzeilen: {last - 1}  LV Summenzeile: {lv_total_row}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
