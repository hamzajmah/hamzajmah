"""T6b Managementbericht als Excel.

Ein Blatt je Frage, die das Management wirklich stellt:

    Wie viel ist geliefert worden?      -> Uebersicht
    Wo ist es hingegangen?              -> Orte
    Was steckt in welchem Bauabschnitt? -> Bereiche
    Wie hat es sich entwickelt?         -> Zeitverlauf
    Woraus besteht es?                  -> Material
    Wie viel davon ist abgerechnet?     -> LV_Abgleich
    Wie sehr kann ich den Zahlen trauen? -> Datenqualitaet, Quellen_Annahmen

Alle Kennzahlen sind Formeln auf das Blatt `Daten`. Wer eine Zahl anzweifelt,
filtert dort und sieht die Zeilen dahinter.
"""
from __future__ import annotations

import csv
import json
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..config import load_lv_mapping
from ..harness import Context, TaskResult
from ..materials import CHARGE_SUPPLY
from ..state import Task

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="D9E2F3")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(bottom=THIN)

DATA_COLUMNS = [
    "ort", "ortstyp", "ortsquelle", "bereich", "bereichsgruppe", "sektion",
    "bauweise", "km_von", "km_bis", "materialgruppe", "material", "monat",
    "fuhren", "menge_t", "m3_lose", "m3_eingebaut", "m3_min", "m3_max", "in_pruefung_t",
]
GROUP_LABELS = {
    "bedding_supplied": "Bettungsmaterial",
    "gravel_base_layer": "Schotter Tragschicht",
    "hdd_platform": "Arbeitsplattform Querung",
    "sand_cover": "Sandabdeckung",
    "nicht_zugeordnet": "nicht zugeordnet",
}
AREA_GROUPS = {
    "section": "Trassenabschnitte",
    "crossing": "Querungen",
    "joint_pit": "Muffen- und Schubgruben",
    "general": "Nicht bereichsscharf",
}


@dataclass
class Bucket:
    fuhren: int = 0
    menge_t: float = 0.0
    m3_lose: float = 0.0
    m3_eingebaut: float = 0.0
    m3_min: float = 0.0
    m3_max: float = 0.0
    pruefung_t: float = 0.0
    sources: set[str] = field(default_factory=set)


def run(task: Task, ctx: Context) -> TaskResult:
    records = [
        r for r in ctx.store.records()
        if r.charge_type == CHARGE_SUPPLY and not r.is_duplicate
    ]
    if not records:
        return TaskResult(ok=False, message="keine Lieferungen", error_class="no_records", escalate=False)

    mapping = load_lv_mapping(ctx.cfg).get("material_to_group") or {}
    structures = _structures(ctx.work_dir)

    buckets: dict[tuple, Bucket] = defaultdict(Bucket)
    for rec in records:
        month = rec.delivery_date.strftime("%Y-%m") if rec.delivery_date else "ohne Datum"
        material = rec.material_key() or "unbekannt"
        group = GROUP_LABELS.get(mapping.get(material, "nicht_zugeordnet"), "nicht zugeordnet")
        area = rec.area_final or "GENERAL"
        location = rec.location_label or "ohne Ortsangabe"
        key = (location, rec.location_type or "none", area, rec.area_class or "unknown", group, material, month)
        b = buckets[key]
        b.fuhren += 1
        b.menge_t += rec.quantity_t or 0.0
        b.m3_lose += rec.delivered_m3_loose or 0.0
        b.m3_eingebaut += rec.delivered_m3_installed or 0.0
        b.m3_min += rec.delivered_m3_installed_low or 0.0
        b.m3_max += rec.delivered_m3_installed_high or 0.0
        b.pruefung_t += (rec.quantity_t or 0.0) if rec.needs_review else 0.0
        if rec.location_source:
            b.sources.add(rec.location_source)

    wb = Workbook()
    wb.remove(wb.active)
    materials = sorted({key[5] for key in buckets}, key=lambda m: -sum(b.menge_t for k, b in buckets.items() if k[5] == m))
    data_rows = _write_data(wb, buckets, structures)
    # Die letzte Datenzeile kommt aus dem Blatt selbst, nicht aus der Listenlaenge.
    # Eine Formel, die eine Zeile zu frueh endet, faellt niemandem auf.
    last = wb["Daten"].max_row
    _write_factors(wb, ctx, materials, last)
    wb.move_sheet("Faktoren", offset=-(len(wb.sheetnames) - 1))

    comparison = _comparison(ctx.work_dir)
    quality = _quality(ctx.work_dir)

    present_groups = sorted({str(row[9]) for row in data_rows})
    _write_overview(wb, ctx, last, comparison, quality, present_groups, materials)
    _write_locations(wb, data_rows, last)
    _write_areas(wb, data_rows, last, comparison)
    _write_timeline(wb, data_rows, last)
    _write_material(wb, ctx, data_rows, last)
    _write_lv(wb, comparison, last)
    _write_quality(wb, ctx, quality, last)
    _write_sources(wb, ctx)

    wb.move_sheet("Daten", offset=len(wb.sheetnames))
    path = ctx.output_dir / "management_bericht.xlsx"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

    ctx.log(f"MANAGEMENT bericht={path.name} datenzeilen={len(data_rows)}")
    return TaskResult(ok=True, message=f"Managementbericht mit {len(data_rows)} Datenzeilen", data={"rows": len(data_rows)})


# -- Hilfen ---------------------------------------------------------------
def _structures(work_dir: Path) -> dict[str, dict[str, str]]:
    path = work_dir / "01_structures.csv"
    out: dict[str, dict[str, str]] = {}
    if not path.exists():
        return out
    with path.open(encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh, delimiter=";"):
            out.setdefault(row["area_key"], row)
    return out


def _comparison(work_dir: Path) -> list[dict]:
    path = work_dir / "06_comparison.csv"
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as fh:
        return [r for r in csv.DictReader(fh, delimiter=";") if r["in_comparison_period"].lower() == "true"]


def _quality(work_dir: Path) -> dict:
    path = work_dir / "quality.json"
    return json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}


def _style_header(ws: Worksheet, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    # Als Zeichenkette setzen: ws.cell(...) wuerde die Zelle anlegen und damit
    # eine leere Zeile erzeugen, die alle folgenden Zeilen verschiebt.
    ws.freeze_panes = f"A{row + 1}"


def _title(ws: Worksheet, text: str, subtitle: str = "") -> int:
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, bold=True, size=14, color="1F3864")
    row = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name=FONT, size=9, italic=True, color="595959")
        row = 3
    return row + 1


def _widths(ws: Worksheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _number(ws: Worksheet, ref: str, fmt: str = "#,##0") -> None:
    ws[ref].number_format = fmt
    ws[ref].font = Font(name=FONT, size=10)


# -- Blaetter -------------------------------------------------------------
def _write_data(wb: Workbook, buckets: dict[tuple, Bucket], structures: dict) -> list[list]:
    ws = wb.create_sheet("Daten")
    ws.append(DATA_COLUMNS)
    _style_header(ws, 1, len(DATA_COLUMNS))

    rows: list[list] = []
    for key in sorted(buckets, key=lambda k: (k[2], k[0], k[4], k[6])):
        location, kind, area, area_class, group, material, month = key
        b = buckets[key]
        structure = structures.get(area, {})
        rows.append([
            location,
            {"point": "Setzpunkt", "crossing": "Querungsbauwerk", "be_area": "BE Flaeche", "none": "ohne Ortsangabe"}.get(kind, kind),
            ", ".join(sorted(b.sources)) or "keine",
            area,
            AREA_GROUPS.get(area_class, "Sonstige"),
            structure.get("section_key", ""),
            structure.get("method", ""),
            float(structure["km_von"]) if structure.get("km_von") else None,
            float(structure["km_bis"]) if structure.get("km_bis") else None,
            group, material, month,
            b.fuhren, round(b.menge_t, 2), None, None, None, None, round(b.pruefung_t, 2),
        ])
    for row in rows:
        ws.append(row)

    # Kubikmeter entstehen erst hier, aus Tonnen geteilt durch die Dichte des
    # jeweiligen Materials. Wer im Blatt Faktoren eine Dichte aendert, sieht die
    # Wirkung sofort in jeder Auswertung.
    lookup = "Faktoren!$A$4:$A$40"
    for index in range(len(rows)):
        r = index + 2
        ws.cell(row=r, column=15, value=f'=IFERROR($N{r}/INDEX(Faktoren!$B$4:$B$40,MATCH($K{r},{lookup},0)),"")')
        ws.cell(row=r, column=16, value=f'=IFERROR($N{r}/INDEX(Faktoren!$C$4:$C$40,MATCH($K{r},{lookup},0)),"")')
        ws.cell(row=r, column=17, value=f'=IFERROR($N{r}/(INDEX(Faktoren!$C$4:$C$40,MATCH($K{r},{lookup},0))*(1+Faktoren!$D$1)),"")')
        ws.cell(row=r, column=18, value=f'=IFERROR($N{r}/(INDEX(Faktoren!$C$4:$C$40,MATCH($K{r},{lookup},0))*(1-Faktoren!$D$1)),"")')
    for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1):
        for cell in row:
            cell.font = Font(name=FONT, size=9)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00" if cell.column >= 14 else "#,##0"
    _widths(ws, {"A": 18, "B": 18, "C": 22, "D": 16, "E": 22, "F": 14, "G": 18, "J": 22, "K": 20, "L": 10})
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DATA_COLUMNS))}{ws.max_row}"
    return rows


def _write_factors(wb: Workbook, ctx: Context, materials: list[str], last: int = 0) -> int:
    """Die Stellschraube des ganzen Berichts, an einer Stelle und aenderbar.

    Jede Kubikmeterangabe im Bericht rechnet gegen diese Tabelle. Eine Dichte
    hier zu aendern aendert Uebersicht, Orte, Bereiche und den LV Abgleich
    zugleich - und macht damit sichtbar, wie stark die Aussage vom Faktor
    abhaengt.
    """
    ws = wb.create_sheet("Faktoren")
    row = _title(
        ws,
        "Umrechnung Tonnen in Kubikmeter",
        "Gelbe Zellen sind aenderbar. Jede Kubikmeterzahl im Bericht rechnet gegen diese Tabelle.",
    )
    sensitivity = float((ctx.factors.get("defaults") or {}).get("sensitivity_pct", 10.0)) / 100.0
    ws["D1"] = sensitivity
    ws["D1"].number_format = "0%"
    ws["D1"].fill = NOTE_FILL
    ws["D1"].font = Font(name=FONT, size=10, bold=True, color="0000FF")
    ws["E1"] = "Sensitivitaet, wirkt auf die Bandbreite"
    ws["E1"].font = Font(name=FONT, size=9, italic=True, color="595959")

    header = [
        "Material", "Dichte lose (t/m3)", "Dichte eingebaut (t/m3)",
        "m3 je t lose", "m3 je t eingebaut", "Quelle des Faktors", "Konfidenz",
    ]
    for index, name in enumerate(header, start=1):
        ws.cell(row=row, column=index, value=name)
    _style_header(ws, row, len(header))

    factors = ctx.factors.get("factors") or {}
    first = row + 1
    for offset, material in enumerate(materials):
        r = first + offset
        entry = factors.get(material.replace(" ", "_").replace("/", "_"), {})
        ws.cell(row=r, column=1, value=material)
        for col, key in ((2, "bulk_density_t_per_m3"), (3, "installed_density_t_per_m3")):
            cell = ws.cell(row=r, column=col, value=entry.get(key))
            cell.number_format = "0.00"
            cell.fill = NOTE_FILL
            cell.font = Font(name=FONT, size=10, bold=True, color="0000FF")
        ws.cell(row=r, column=4, value=f'=IFERROR(1/B{r},"")').number_format = "0.000"
        ws.cell(row=r, column=5, value=f'=IFERROR(1/C{r},"")').number_format = "0.000"
        ws.cell(row=r, column=6, value=entry.get("source", "kein Faktor hinterlegt"))
        ws.cell(row=r, column=7, value=entry.get("confidence", "-"))
        for col in (1, 4, 5, 6, 7):
            ws.cell(row=r, column=col).font = Font(name=FONT, size=10)
        if entry.get("confidence") == "assumption":
            ws.cell(row=r, column=7).fill = NOTE_FILL

    # Wirkung sichtbar machen: was ein pauschaler Faktor aus derselben Menge macht.
    compare_row = first + len(materials) + 1
    ws.cell(row=compare_row, column=1, value="Was ein pauschaler Faktor daraus machen wuerde").font = Font(
        name=FONT, bold=True, size=11, color="1F3864"
    )
    compare_row += 1
    for index, name in enumerate(["Vergleich", "m3 eingebaut", "abgerechnet laut LV", "Deckungsgrad", "Bemerkung"], start=1):
        ws.cell(row=compare_row, column=index, value=name)
    _style_header(ws, compare_row, 5)

    billed_all = "SUM(LV_Abgleich!$D$4:$D$400)/2"
    lines = [
        (
            "materialspezifisch, wie hinterlegt",
            f"=SUM(Daten!$P$2:$P${last})" if last else "",
            "gerechnet je Material mit eigener Dichte",
        ),
        (
            "pauschal 0,4 m3 je t",
            f"=SUM(Daten!$N$2:$N${last})*0.4" if last else "",
            "entspricht 2,5 t je m3 fuer alles",
        ),
        (
            "pauschal 0,5 m3 je t",
            f"=SUM(Daten!$N$2:$N${last})*0.5" if last else "",
            "entspricht 2,0 t je m3 fuer alles",
        ),
    ]
    start_compare = compare_row + 1
    for offset, (label, formula, remark) in enumerate(lines):
        r = start_compare + offset
        ws.cell(row=r, column=1, value=label).font = Font(name=FONT, size=10, bold=offset == 0)
        cell = ws.cell(row=r, column=2, value=formula)
        cell.number_format = "#,##0"
        cell.font = Font(name=FONT, size=10, bold=offset == 0)
        billed = ws.cell(row=r, column=3, value=f"={billed_all}")
        billed.number_format = "#,##0"
        billed.font = Font(name=FONT, size=10)
        ratio = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/B{r},0)")
        ratio.number_format = "0.0%"
        ratio.font = Font(name=FONT, size=10, bold=offset == 0)
        ws.cell(row=r, column=5, value=remark).font = Font(name=FONT, size=9, color="595959")

    hint = ws.cell(
        row=start_compare + len(lines) + 1,
        column=1,
        value=(
            "Ein Deckungsgrad ueber 100 Prozent bedeutet, dass mehr abgerechnet als geliefert waere. Beim "
            "Bettungsmaterial ist das ein Warnsignal fuer einen zu hohen Dichtewert."
        ),
    )
    hint.font = Font(name=FONT, size=9, italic=True)
    hint.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=hint.row, start_column=1, end_row=hint.row + 1, end_column=5)

    note_row = hint.row + 3
    note = ws.cell(
        row=note_row,
        column=1,
        value=(
            "Blau geschriebene Zellen sind Eingaben, schwarze sind gerechnet. Zum Vergleich: ein pauschaler Faktor "
            "von 0,4 m3 je t entspricht einer Dichte von 2,5 t je m3 und liegt damit ueber jeder Einbaudichte eines "
            "Gemisches. Wird er gesetzt, faellt das ausgewiesene Volumen um rund ein Viertel."
        ),
    )
    note.font = Font(name=FONT, size=9, italic=True)
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=7)
    _widths(ws, {"A": 24, "B": 20, "C": 24, "D": 16, "E": 20, "F": 54, "G": 14})
    return first + len(materials) - 1


def _write_overview(
    wb: Workbook,
    ctx: Context,
    last: int,
    comparison: list[dict],
    quality: dict,
    present_groups: list[str],
    present_materials: list[str],
) -> None:
    ws = wb.create_sheet("Uebersicht")
    period_from, period_to = (d.isoformat() for d in ctx.cfg.period)
    row = _title(
        ws,
        f"Schotter und Bettungsmaterial - {ctx.cfg['project']['name']}",
        f"Betrachtungszeitraum {period_from} bis {period_to}. Alle Kennzahlen sind Formeln auf das Blatt Daten.",
    )

    blocks: list[tuple[str, list[tuple[str, object, str, str]]]] = [
        ("Wie viel ist geliefert worden", [
            ("Liefermenge", f"=SUM(Daten!N2:N{last})", "#,##0", "t"),
            ("Fuhren", f"=SUM(Daten!M2:M{last})", "#,##0", "Anzahl"),
            ("Volumen eingebaut", f"=SUM(Daten!P2:P{last})", "#,##0", "m3"),
            ("Volumen eingebaut, untere Grenze", f"=SUM(Daten!Q2:Q{last})", "#,##0", "m3 bei Dichte plus 10 Prozent"),
            ("Volumen eingebaut, obere Grenze", f"=SUM(Daten!R2:R{last})", "#,##0", "m3 bei Dichte minus 10 Prozent"),
            ("Volumen lose, zum Vergleich", f"=SUM(Daten!O2:O{last})", "#,##0", "m3 bei Anlieferung"),
        ]),
        ("Wovon, in Tonnen", [
            (label, f'=SUMIFS(Daten!N2:N{last},Daten!J2:J{last},"{label}")', "#,##0", "t")
            for label in present_groups
        ]),
        ("Wovon, in Kubikmetern eingebaut, je Material mit eigenem Faktor", [
            (
                material,
                f'=SUMIFS(Daten!P2:P{last},Daten!K2:K{last},"{material}")',
                "#,##0",
                f'=IFERROR("Dichte "&TEXT(INDEX(Faktoren!$C$4:$C$40,MATCH("{material}",Faktoren!$A$4:$A$40,0)),"0.00")&" t je m3","")',
            )
            for material in present_materials
        ]),
        ("Wo", [
            ("Trassenabschnitte", f'=SUMIFS(Daten!N2:N{last},Daten!E2:E{last},"Trassenabschnitte")', "#,##0", "t"),
            ("Querungen", f'=SUMIFS(Daten!N2:N{last},Daten!E2:E{last},"Querungen")', "#,##0", "t"),
            ("Muffen- und Schubgruben", f'=SUMIFS(Daten!N2:N{last},Daten!E2:E{last},"Muffen- und Schubgruben")', "#,##0", "t"),
            ("Nicht bereichsscharf gebucht", f'=SUMIFS(Daten!N2:N{last},Daten!E2:E{last},"Nicht bereichsscharf")', "#,##0", "t"),
            ("Menge mit Ortsangabe", f'=SUMIFS(Daten!N2:N{last},Daten!B2:B{last},"<>ohne Ortsangabe")', "#,##0", "t"),
            ("Anteil mit Ortsangabe", f'=IFERROR(SUMIFS(Daten!N2:N{last},Daten!B2:B{last},"<>ohne Ortsangabe")/SUM(Daten!N2:N{last}),0)', "0.0%", "der Liefermenge"),
        ]),
        ("Wie viel davon ist abgerechnet", [
            ("Bettungsmaterial, geliefert", '=IFERROR(INDEX(LV_Abgleich!$C$4:$C$400,MATCH("Bettungsmaterial",LV_Abgleich!$A$4:$A$400,0)),0)', "#,##0", "m3 eingebaut"),
            ("Bettungsmaterial, abgerechnet", '=IFERROR(INDEX(LV_Abgleich!$D$4:$D$400,MATCH("Bettungsmaterial",LV_Abgleich!$A$4:$A$400,0)),0)', "#,##0", "m3 laut Leistungsmeldung"),
            ("Schotter, geliefert", '=IFERROR(INDEX(LV_Abgleich!$C$4:$C$400,MATCH("Schotter Tragschicht",LV_Abgleich!$A$4:$A$400,0)),0)', "#,##0", "m3 eingebaut"),
            ("Schotter, abgerechnet", '=IFERROR(INDEX(LV_Abgleich!$D$4:$D$400,MATCH("Schotter Tragschicht",LV_Abgleich!$A$4:$A$400,0)),0)', "#,##0", "m3 laut Leistungsmeldung"),
        ]),
        ("Wie sehr kann ich den Zahlen trauen", [
            ("Menge in Pruefung", f"=SUM(Daten!S2:S{last})", "#,##0", "t"),
            ("Anteil in Pruefung", f"=IFERROR(SUM(Daten!S2:S{last})/SUM(Daten!N2:N{last}),0)", "0.0%", f"Schwelle {quality.get('review_share_threshold_pct', 5)} Prozent"),
            ("Offene Entscheidungen", len(ctx.decisions.items()), "#,##0", "siehe DECISIONS.md"),
        ]),
    ]

    for heading, entries in blocks:
        ws.cell(row=row, column=1, value=heading).font = Font(name=FONT, bold=True, size=11, color="1F3864")
        ws.cell(row=row, column=1).fill = SUB_FILL
        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = SUB_FILL
        row += 1
        for label, value, fmt, unit in entries:
            ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10)
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = fmt
            cell.font = Font(name=FONT, size=10, bold=True)
            unit_cell = ws.cell(row=row, column=3, value=unit)
            unit_cell.font = Font(name=FONT, size=9, color="595959")
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = BORDER
            row += 1
        row += 1

    note = ws.cell(
        row=row,
        column=1,
        value=(
            "Kubikmeter sind umgerechnet, nicht gemessen. Die Umrechnung je Material steht im Blatt Material "
            "samt Quelle. Der Vergleich gegen das Leistungsverzeichnis erfolgt ausschliesslich gegen das "
            "verdichtete Einbauvolumen."
        ),
    )
    note.font = Font(name=FONT, size=9, italic=True)
    note.fill = NOTE_FILL
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=4)
    _widths(ws, {"A": 38, "B": 18, "C": 34, "D": 20})


def _write_locations(wb: Workbook, rows: list[list], last: int) -> None:
    ws = wb.create_sheet("Orte")
    row = _title(ws, "Wo ist das Material angekommen", "Je Ort, sortiert nach Menge. Werte sind Formeln auf das Blatt Daten.")
    header = ["Ort", "Art", "Bereich", "Sektion", "Bauweise", "Fuhren", "Menge (t)", "Bettung (t)", "Schotter (t)", "m3 eingebaut", "Anteil an Gesamtmenge"]
    ws.append([])  # Platzhalter, damit die Kopfzeile in der berechneten Zeile steht
    for index, name in enumerate(header, start=1):
        ws.cell(row=row, column=index, value=name)
    _style_header(ws, row, len(header))

    seen: dict[str, list] = {}
    for entry in rows:
        seen.setdefault(entry[0], entry)
    order = sorted(seen, key=lambda name: -sum(e[13] for e in rows if e[0] == name))

    first_data = row + 1
    for offset, location in enumerate(order):
        r = first_data + offset
        sample = seen[location]
        ws.cell(row=r, column=1, value=location)
        ws.cell(row=r, column=2, value=sample[1])
        ws.cell(row=r, column=3, value=sample[3])
        ws.cell(row=r, column=4, value=sample[5])
        ws.cell(row=r, column=5, value=sample[6])
        ws.cell(row=r, column=6, value=f'=SUMIFS(Daten!M2:M{last},Daten!A2:A{last},$A{r})')
        ws.cell(row=r, column=7, value=f'=SUMIFS(Daten!N2:N{last},Daten!A2:A{last},$A{r})')
        ws.cell(row=r, column=8, value=f'=SUMIFS(Daten!N2:N{last},Daten!A2:A{last},$A{r},Daten!J2:J{last},"Bettungsmaterial")')
        ws.cell(row=r, column=9, value=f'=SUMIFS(Daten!N2:N{last},Daten!A2:A{last},$A{r},Daten!J2:J{last},"Schotter Tragschicht")')
        ws.cell(row=r, column=10, value=f'=SUMIFS(Daten!P2:P{last},Daten!A2:A{last},$A{r})')
        ws.cell(row=r, column=11, value=f'=IFERROR($G{r}/SUM(Daten!N2:N{last}),0)')
        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name=FONT, size=10)
            if col in (6, 7, 8, 9, 10):
                cell.number_format = "#,##0"
            if col == 11:
                cell.number_format = "0.0%"

    total = first_data + len(order)
    ws.cell(row=total, column=1, value="Summe").font = Font(name=FONT, bold=True, size=10)
    for col, letter in ((6, "F"), (7, "G"), (8, "H"), (9, "I"), (10, "J")):
        cell = ws.cell(row=total, column=col, value=f"=SUM({letter}{first_data}:{letter}{total - 1})")
        cell.number_format = "#,##0"
        cell.font = Font(name=FONT, bold=True, size=10)
    _widths(ws, {"A": 20, "B": 20, "C": 16, "D": 14, "E": 20, "G": 14, "H": 14, "I": 14, "J": 16, "K": 20})


def _write_areas(wb: Workbook, rows: list[list], last: int, comparison: list[dict]) -> None:
    ws = wb.create_sheet("Bereiche")
    row = _title(ws, "Was steckt in welchem Bauabschnitt", "Bereich nach ERP Buchung, Kilometrierung aus den Trassenstammdaten.")
    header = ["Bereich", "Gruppe", "Bauweise", "km von", "km bis", "Fuhren", "Menge (t)", "m3 eingebaut", "Orte im Bereich"]
    for index, name in enumerate(header, start=1):
        ws.cell(row=row, column=index, value=name)
    _style_header(ws, row, len(header))

    areas: dict[str, list] = {}
    for entry in rows:
        areas.setdefault(entry[3], entry)
    order = sorted(areas, key=lambda name: -sum(e[13] for e in rows if e[3] == name))

    first = row + 1
    for offset, area in enumerate(order):
        r = first + offset
        sample = areas[area]
        locations = sorted({e[0] for e in rows if e[3] == area})
        ws.cell(row=r, column=1, value=area)
        ws.cell(row=r, column=2, value=sample[4])
        ws.cell(row=r, column=3, value=sample[6])
        ws.cell(row=r, column=4, value=sample[7])
        ws.cell(row=r, column=5, value=sample[8])
        ws.cell(row=r, column=6, value=f'=SUMIFS(Daten!M2:M{last},Daten!D2:D{last},$A{r})')
        ws.cell(row=r, column=7, value=f'=SUMIFS(Daten!N2:N{last},Daten!D2:D{last},$A{r})')
        ws.cell(row=r, column=8, value=f'=SUMIFS(Daten!P2:P{last},Daten!D2:D{last},$A{r})')
        ws.cell(row=r, column=9, value=", ".join(locations[:6]) + (" ..." if len(locations) > 6 else ""))
        for col in range(1, 10):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name=FONT, size=10)
            if col in (4, 5, 6, 7, 8):
                cell.number_format = "#,##0"
    total = first + len(order)
    ws.cell(row=total, column=1, value="Summe").font = Font(name=FONT, bold=True, size=10)
    for col, letter in ((6, "F"), (7, "G"), (8, "H")):
        cell = ws.cell(row=total, column=col, value=f"=SUM({letter}{first}:{letter}{total - 1})")
        cell.number_format = "#,##0"
        cell.font = Font(name=FONT, bold=True, size=10)
    _widths(ws, {"A": 18, "B": 24, "C": 20, "F": 12, "G": 14, "H": 16, "I": 46})


def _write_timeline(wb: Workbook, rows: list[list], last: int) -> None:
    ws = wb.create_sheet("Zeitverlauf")
    row = _title(ws, "Wie hat sich die Liefermenge entwickelt", "Monatswerte. Buchungsdatum aus dem ERP, siehe Hinweis im Blatt Datenqualitaet.")
    header = ["Monat", "Fuhren", "Menge (t)", "m3 eingebaut", "Menge kumuliert (t)"]
    for index, name in enumerate(header, start=1):
        ws.cell(row=row, column=index, value=name)
    _style_header(ws, row, len(header))

    months = sorted({entry[11] for entry in rows})
    first = row + 1
    for offset, month in enumerate(months):
        r = first + offset
        ws.cell(row=r, column=1, value=month)
        ws.cell(row=r, column=2, value=f'=SUMIFS(Daten!M2:M{last},Daten!L2:L{last},$A{r})')
        ws.cell(row=r, column=3, value=f'=SUMIFS(Daten!N2:N{last},Daten!L2:L{last},$A{r})')
        ws.cell(row=r, column=4, value=f'=SUMIFS(Daten!P2:P{last},Daten!L2:L{last},$A{r})')
        ws.cell(row=r, column=5, value=f"=SUM($C${first}:$C{r})")
        for col in range(1, 6):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name=FONT, size=10)
            if col > 1:
                cell.number_format = "#,##0"
    _widths(ws, {"A": 14, "B": 12, "C": 14, "D": 16, "E": 20})


def _write_material(wb: Workbook, ctx: Context, rows: list[list], last: int) -> None:
    ws = wb.create_sheet("Material")
    row = _title(ws, "Woraus besteht die Menge", "Die Umrechnung in Kubikmeter je Material, mit Quelle und Konfidenz.")
    header = ["Material", "Verwendung", "Menge (t)", "m3 lose", "m3 eingebaut", "Dichte lose (t/m3)", "Dichte eingebaut (t/m3)", "Quelle des Faktors", "Konfidenz"]
    for index, name in enumerate(header, start=1):
        ws.cell(row=row, column=index, value=name)
    _style_header(ws, row, len(header))

    factors = ctx.factors.get("factors") or {}
    materials: dict[str, str] = {}
    for entry in rows:
        materials.setdefault(entry[10], entry[9])

    first = row + 1
    for offset, material in enumerate(sorted(materials, key=lambda m: -sum(e[13] for e in rows if e[10] == m))):
        r = first + offset
        key = material.replace(" ", "_").replace("/", "_")
        entry = factors.get(key, {})
        ws.cell(row=r, column=1, value=material)
        ws.cell(row=r, column=2, value=materials[material])
        ws.cell(row=r, column=3, value=f'=SUMIFS(Daten!N2:N{last},Daten!K2:K{last},$A{r})')
        ws.cell(row=r, column=4, value=f'=SUMIFS(Daten!O2:O{last},Daten!K2:K{last},$A{r})')
        ws.cell(row=r, column=5, value=f'=SUMIFS(Daten!P2:P{last},Daten!K2:K{last},$A{r})')
        ws.cell(row=r, column=6, value=f'=IFERROR(INDEX(Faktoren!$B$4:$B$40,MATCH($A{r},Faktoren!$A$4:$A$40,0)),"")')
        ws.cell(row=r, column=7, value=f'=IFERROR(INDEX(Faktoren!$C$4:$C$40,MATCH($A{r},Faktoren!$A$4:$A$40,0)),"")')
        ws.cell(row=r, column=8, value=entry.get("source", "kein Faktor hinterlegt"))
        ws.cell(row=r, column=9, value=entry.get("confidence", "-"))
        for col in range(1, 10):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(name=FONT, size=10)
            if col in (3, 4, 5):
                cell.number_format = "#,##0"
            if col in (6, 7):
                cell.number_format = "0.00"
        if entry.get("confidence") == "assumption":
            ws.cell(row=r, column=9).fill = NOTE_FILL
    _widths(ws, {"A": 24, "B": 24, "C": 14, "D": 14, "E": 16, "F": 18, "G": 20, "H": 52, "I": 14})


def _write_lv(wb: Workbook, comparison: list[dict], last: int) -> None:
    ws = wb.create_sheet("LV_Abgleich")
    row = _title(
        ws,
        "Wie viel davon ist abgerechnet",
        "Vergleich ausschliesslich gegen das verdichtete Einbauvolumen und nur im Zeitraum, in dem beide Seiten Daten haben.",
    )
    header = ["Materialgruppe", "Bereich", "m3 eingebaut geliefert", "m3 abgerechnet", "Delta m3", "Deckungsgrad", "Delta bei Dichte -10%", "Delta bei Dichte +10%"]
    for index, name in enumerate(header, start=1):
        ws.cell(row=row, column=index, value=name)
    _style_header(ws, row, len(header))

    totals: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0, 0.0, 0.0])
    by_area: dict[tuple[str, str], list[float]] = defaultdict(lambda: [0.0, 0.0, 0.0, 0.0])
    for entry in comparison:
        group = GROUP_LABELS.get(entry["material_group"], entry["material_group"])
        values = [
            float(entry["delivered_m3_installed"] or 0),
            float(entry["billed_m3"] or 0),
            float(entry["delivered_m3_installed_low"] or 0),
            float(entry["delivered_m3_installed_high"] or 0),
        ]
        for index, value in enumerate(values):
            totals[group][index] += value
            by_area[(group, entry["area_final"])][index] += value

    current = row + 1
    for group in sorted(totals, key=lambda g: -totals[g][0]):
        installed, billed, low, high = totals[group]
        _lv_line(ws, current, group, "alle Bereiche", billed, last, bold=True)
        current += 1
        for (g, area) in sorted([k for k in by_area if k[0] == group], key=lambda k: -by_area[k][0]):
            area_values = by_area[(g, area)]
            if area_values[0] == 0 and area_values[1] == 0:
                continue
            _lv_line(ws, current, g, area, area_values[1], last)
            current += 1
        current += 1
    _widths(ws, {"A": 26, "B": 18, "C": 22, "D": 18, "E": 14, "F": 16, "G": 22, "H": 22})


def _lv_line(ws: Worksheet, row: int, group: str, area: str, billed: float, last: int, bold: bool = False) -> None:
    """Geliefert kommt aus dem Datenblatt, abgerechnet aus der Leistungsmeldung.

    Die Lieferseite ist bewusst eine Formel: aendert sich eine Dichte im Blatt
    Faktoren, aendert sich hier der Deckungsgrad mit.
    """
    ws.cell(row=row, column=1, value=group)
    ws.cell(row=row, column=2, value=area)
    scope = (
        f'Daten!$J$2:$J${last},$A{row}'
        if area == "alle Bereiche"
        else f'Daten!$J$2:$J${last},$A{row},Daten!$D$2:$D${last},$B{row}'
    )
    ws.cell(row=row, column=3, value=f"=SUMIFS(Daten!$P$2:$P${last},{scope})")
    ws.cell(row=row, column=4, value=round(billed, 2))
    ws.cell(row=row, column=5, value=f"=D{row}-C{row}")
    ws.cell(row=row, column=6, value=f"=IFERROR(D{row}/C{row},0)")
    ws.cell(row=row, column=7, value=f"=D{row}-SUMIFS(Daten!$R$2:$R${last},{scope})")
    ws.cell(row=row, column=8, value=f"=D{row}-SUMIFS(Daten!$Q$2:$Q${last},{scope})")
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT, size=10, bold=bold)
        if col in (3, 4, 5, 7, 8):
            cell.number_format = "#,##0"
        if col == 6:
            cell.number_format = "0.0%"
        if bold:
            cell.fill = SUB_FILL


def _write_quality(wb: Workbook, ctx: Context, quality: dict, last: int) -> None:
    ws = wb.create_sheet("Datenqualitaet")
    row = _title(ws, "Wie sehr kann ich den Zahlen trauen", "Diese Seite ist der Grund, warum die anderen glaubwuerdig sind.")
    entries = [
        ("Erfasste Lieferpositionen", quality.get("records_supply", 0), "Anzahl"),
        ("Liefermenge gesamt", quality.get("supply_t_total", 0), "t"),
        ("Menge in der Pruefliste", quality.get("supply_t_in_review", 0), "t"),
        ("Anteil in der Pruefliste", (quality.get("review_share_pct", 0) or 0) / 100.0, "Schwelle 5 Prozent"),
        ("Als Dublette ausgesteuert", quality.get("duplicates_removed", 0), "Zeilen"),
        ("Ergebnis vorlaeufig", "ja" if quality.get("result_provisional") else "nein", "wenn Pruefliste ueber Schwelle"),
        ("Menge ohne Ortsangabe", f'=SUMIFS(Daten!N2:N{last},Daten!B2:B{last},"ohne Ortsangabe")', "t"),
        ("Offene Entscheidungen", len(ctx.decisions.items()), "siehe DECISIONS.md"),
    ]
    for label, value, unit in entries:
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10)
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.number_format = "0.0%" if "Anteil" in label else "#,##0"
        ws.cell(row=row, column=3, value=unit).font = Font(name=FONT, size=9, color="595959")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Was diese Auswertung nicht zeigt").font = Font(name=FONT, bold=True, size=11, color="1F3864")
    row += 1
    for text in (
        "Kubikmeter sind gerechnet, nicht gemessen. Die Dichten sind Literaturwerte ohne Projektbeleg.",
        "Ein Teil der Wareneingaenge traegt als Lieferdatum den Monatsersten. Monatssummen sind belastbar, Tagesverlaeufe nicht.",
        "Zuschlaege und Annahmemengen der Bestellung P100012091 fehlen, weil dafuer kein Wareneingangsexport vorliegt.",
        "Aussagen zu Verlusten oder Mehrverbrauch sind erst nach Abarbeitung der Pruefliste belastbar.",
    ):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name=FONT, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
    _widths(ws, {"A": 46, "B": 16, "C": 34, "D": 20})


def _write_sources(wb: Workbook, ctx: Context) -> None:
    ws = wb.create_sheet("Quellen_Annahmen")
    row = _title(ws, "Woher die Zahlen stammen", "Jede Quelle mit ihrer Rolle. Gelbe Zellen sind Annahmen, keine Messwerte.")
    header = ["Quelle", "Rolle", "Was daraus kommt"]
    for index, name in enumerate(header, start=1):
        ws.cell(row=row, column=index, value=name)
    _style_header(ws, row, len(header))
    row += 1

    sources = [
        ("IFS Wareneingang P100042563", "fuehrend", "Mengen, Bereiche, Ortsnotizen ab Januar 2026"),
        ("Tracking Mappe, Blatt Lieferungen", "fuehrend fuer P100012091", "Mengen 2024 und 2025, Ortszuordnung aus Originalbelegen"),
        ("Leistungsmeldung", "fuehrend fuer die Abrechnung", "Vertragsmengen, abgerechnete Mengen, Erloese"),
        ("Trassenstammdaten Bauwerksflaechen", "Stammdaten", "Sektion, Kilometrierung, Bauweise je Querung"),
        ("Lieferlog des Lieferanten", "Gegenprobe", "Kontrolle der Menge, zweite Ortsquelle"),
        ("IFS Rechnungsliste", "Gegenprobe", "Bestellungen und Betraege"),
    ]
    for name, role, content in sources:
        for col, value in enumerate((name, role, content), start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Annahmen").font = Font(name=FONT, bold=True, size=11, color="1F3864")
    row += 1
    for text in (
        "Dichten je Material aus Literaturbandbreiten, kein Projektbeleg. Sensitivitaet plus minus 10 Prozent.",
        "Zuordnung Material zu LV Position laut Bauleitung: 0/8 und 0/22 Bettung, 0/45 und 50/200 Schotter, 0/2 Muffengruben.",
        "LV Kapitel ohne Bereichsangabe werden dem Sammelbereich GENERAL zugeordnet.",
        "Zuschlagszeilen tragen Tonnen nur als Abrechnungsbasis und zaehlen nie als Liefermenge.",
    ):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name=FONT, size=9)
        cell.fill = NOTE_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
    _widths(ws, {"A": 46, "B": 28, "C": 60})
