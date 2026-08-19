"""Excel Ausgabe fuer Power BI.

Formatvorgaben aus Abschnitt 7: je Blatt genau eine Excel Tabelle (ListObject)
mit stabilem Namen gleich dem Blattnamen, Kopfzeile in Zeile 1, keine
Titelzeilen, keine verbundenen Zellen, keine Leerzeilen, keine Summenzeilen,
Zahlen als echte Zahlen, Datumswerte als echte Datumswerte.
"""
from __future__ import annotations

from collections.abc import Mapping, Sequence
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)          # konsistenter Datentyp ueber alle Zeilen
    if isinstance(value, (int, float, datetime, date, str)):
        return value
    return str(value)


def write_workbook(path: Path, tables: Mapping[str, tuple[Sequence[str], Sequence[Sequence[Any]]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for name in tables:
        columns, rows = tables[name]
        ws = wb.create_sheet(title=name)
        ws.append(list(columns))
        for row in rows:
            ws.append([_cell_value(v) for v in row])

        last_col = get_column_letter(len(columns))
        # Eine Tabelle braucht mindestens eine Datenzeile. Leere Faktentabellen
        # bekommen deshalb eine leere Zeile statt erfundener Inhalte.
        last_row = max(len(rows) + 1, 2)
        table = Table(displayName=name, ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(table)

        for idx, column in enumerate(columns, 1):
            width = max(len(str(column)) + 2, 12)
            ws.column_dimensions[get_column_letter(idx)].width = min(width, 40)
            if column.endswith("_date") or column == "date":
                for cell in ws[get_column_letter(idx)][1:]:
                    cell.number_format = "yyyy-mm-dd"
        ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def verify_workbook(path: Path, expected: Mapping[str, Sequence[str]]) -> list[str]:
    """Abnahmepruefung der Arbeitsmappe. Leere Liste = bestanden."""
    from openpyxl import load_workbook

    problems: list[str] = []
    # Ein Durchlauf im Lesemodus: Kopfzeile und Datentypen je Spalte.
    wb = load_workbook(path, read_only=True)
    for name, columns in expected.items():
        if name not in wb.sheetnames:
            problems.append(f"blatt_fehlt:{name}")
            continue
        rows = wb[name].iter_rows(values_only=True)
        header = list(next(rows, ()) or ())[: len(columns)]
        if header != list(columns):
            problems.append(f"kopfzeile_abweichend:{name}")
        kinds: list[set[str]] = [set() for _ in columns]
        for row in rows:
            for index in range(min(len(columns), len(row))):
                value = row[index]
                if value is None or value == "":
                    continue
                kinds[index].add(
                    "date" if isinstance(value, (date, datetime))
                    else ("number" if isinstance(value, (int, float)) else "text")
                )
        for index, found in enumerate(kinds):
            if len(found) > 1:
                problems.append(f"gemischte_datentypen:{name}.{columns[index]}")
    wb.close()

    # Tabellenobjekte und verbundene Zellen brauchen die vollstaendige Mappe.
    wb = load_workbook(path)
    for name in expected:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        tables = list(ws.tables.values())
        if len(tables) != 1:
            problems.append(f"genau_eine_tabelle_erwartet:{name}")
            continue
        if tables[0].displayName != name:
            problems.append(f"tabellenname_ungleich_blattname:{name}")
        if ws.merged_cells.ranges:
            problems.append(f"verbundene_zellen:{name}")
    return problems
