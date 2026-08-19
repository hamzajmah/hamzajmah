"""Testfixtures.

Der ERP Fixture enthaelt echte Positionsbezeichnungen und Notizformate aus dem
Wareneingangsexport des Lieferanten, gekuerzt auf wenige Zeilen.
"""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# Echte Positionsarten des Lieferanten Baustoff Vertrieb Fulda Werra.
ERP_ROWS = [
    # (part description, qty, unit, date, notes, sub project, activity, receipt no, receipt ref, po line)
    ("Bau Tandemzug + Maut : 310080 Mineralgemisch 0/8 (Kalkstein)", 24.65, "ton", "2026-07-22", "SP37 LS 10044", "AS04S-3-04", "Backfill", 243, "D-2604010417", 1),
    ("gleitenden Diesel- /Energie-zuschlag €0,20/t", 24.65, "ton", "2026-07-22", "SP37 LS 10044", "GENERAL", "Backfill", 1705, "D-2604010417", 2),
    ("gleitenden Diesel- /Energie-zuschlag €0,20/t", 24.65, "ton", "2026-07-22", "SP37 LS 10044", "GENERAL", "Backfill", 1722, "D-2604010417", 4),
    ("Sattel/Tandemzug + Maut : Mineralgemisch 50/200 (Porphyr)", 25.10, "ton", "2026-07-23", "Q127 LS 11627", "QR - 127", "Accesses, Parking Areas And Crossings", 244, "D-2604010418", 6),
    ("Bau Tandemzug + Maut : 0/45 mm Mineralgemisch FSS", 24.00, "ton", "2026-07-24", "SP78 as well", "GENERAL", "Trenching", 245, "D-2604010419", 7),
    ("Bau Tandemzug + Maut : Annahme Erdaushub gem. BM-0 mit BGA", 27.42, "ton", "2026-07-24", "SP78 LS 2643503457", "AS04S-3-01", "Slurry And Soil", 246, "D-2604010420", 8),
    ("SP - Samstagszuschlag", 25.00, "ton", "2026-07-25", "SP47 LS 10171", "GENERAL", "Backfill", 671, "D-2604010448", 10),
    ("SP - Frachtkostenausgleich Schüttgut", 12.50, "€", "2026-07-25", "", "GENERAL", "Backfill", 672, "D-2604010448", 12),
    ("Bau Tandemzug + Maut : 0/45 mm Mineralgemisch FSS", 20.00, "ton", "2026-07-26", "SP 48 - SP 51", "GENERAL", "Trenching", 673, "D-2604010449", 14),
    ("Sattel/Tandemzug + Maut : Mineralgemisch 50/200 (Porphyr)", 22.00, "ton", "2026-07-26", "", "GENERAL", "Trenching", 674, "D-2604010450", 16),
]

ERP_COLUMNS = [
    "Source Ref 1", "Source Ref 2", "Source Ref 3", "Receipt No", "Source Ref Type",
    "Sender Description", "Site", "Source Part Description", "Status", "Arrived Source Qty",
    "Measure Unit", "Actual Delivery Date", "Received By", "Receipt Reference", "Notes",
    "Program Description", "Project ID", "Sub Project ID", "Sub Project Description",
    "Activity ID", "Activity Description",
]


def write_erp_fixture(path: Path) -> None:
    import pandas as pd

    rows = []
    for part, qty, unit, day, notes, sub, activity, receipt_no, receipt_ref, po_line in ERP_ROWS:
        rows.append({
            "Source Ref 1": "P100042563", "Source Ref 2": po_line, "Source Ref 3": 1,
            "Receipt No": receipt_no, "Source Ref Type": "PurchaseOrder",
            "Sender Description": "BAUSTOFF VERTRIEB FULDA WERRA GMBH (LEISTUNGSGEMEINSCHAFT SUEDLINK, LOS 4)",
            "Site": "1048A", "Source Part Description": part, "Status": "Received",
            "Arrived Source Qty": qty, "Measure Unit": unit,
            "Actual Delivery Date": datetime.fromisoformat(day),
            "Received By": "TESTUSER", "Receipt Reference": receipt_ref, "Notes": notes,
            "Program Description": "24 DE TRANSNETBW Suedlink Eschwege", "Project ID": 1,
            "Sub Project ID": sub, "Sub Project Description": sub,
            "Activity ID": "201-70000", "Activity Description": activity,
        })
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows, columns=ERP_COLUMNS).to_excel(path, index=False)


# Nachgebildete Leistungsmeldung: gleicher Aufbau wie die echte Datei
# (Kopfzeilen ueber der Tabelle, Gliederungszeilen ohne KT, Wochenspalten in Euro).
LV_HEADER_ROW = 13
LV_ROWS = [
    # (OZ, KT, KURZTEXT, LV-MENGE, EINHEIT, EP, GP, RE MENGE, GP2, KW-Werte)
    ("2", None, "Bauvorbereitende Massnahmen", None, None, None, None, None, None, {}),
    ("2.3", None, "Baustraßen und Zufahrten", None, None, None, None, None, None, {}),
    ("2.3.6", None, "Innere Baustraße", None, None, None, None, None, None, {}),
    ("2.3.6.30", "Haupt-LV", "Schottertragschicht herstellen, 30-40 cm", 42615.0, "m3", 59.35, 2529200.0, 1000.0, 59350.0,
     {"KW 30/26": 59350.0}),
    ("4", None, "Grabentiefbau", None, None, None, None, None, None, {}),
    ("4.4", None, "Abschnitt 04S-3-04", None, None, None, None, None, None, {}),
    ("4.4.4", None, "Bettung", None, None, None, None, None, None, {}),
    ("4.4.4.20", "Haupt-LV", "Bettungsmaterial liefern und einbauen (Leitungszone)", 3000.0, "m3", 51.59, 154770.0, 200.0, 10318.0,
     {"KW 30/26": 10318.0}),
    ("4.4.4.10", "Haupt-LV", "Aufbereitetes Bettungsmaterial einbauen", 1500.0, "m3", 19.46, 29190.0, 0.0, 0.0, {}),
]
LV_WEEK_COLUMNS = ["Bis KW 31/25", "KW 30/26"]


def write_lv_fixture(path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Haupt_Leistung "
    ws["B2"] = "HAUPTVERTRAG"
    ws["B3"] = "PROJEKTNAME:"
    ws["C3"] = "SUEDLINK - BAULOS 4"
    header = ["", "OZ", "KT", "KURZTEXT", "LV-MENGE", "EINHEIT", "EP", "GP", "RE MENGE", "GP", "% Fortschritt", *LV_WEEK_COLUMNS]
    for column, value in enumerate(header, 1):
        ws.cell(row=LV_HEADER_ROW, column=column, value=value)

    row_no = LV_HEADER_ROW + 1
    for oz, kt, text, lv_qty, unit, ep, gp, re_qty, gp2, weeks in LV_ROWS:
        values = ["", oz, kt, text, lv_qty, unit, ep, gp, re_qty, gp2, None]
        for week in LV_WEEK_COLUMNS:
            values.append(weeks.get(week))
        for column, value in enumerate(values, 1):
            ws.cell(row=row_no, column=column, value=value)
        row_no += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


@pytest.fixture()
def project(tmp_path: Path) -> Path:
    """Vollstaendiges Miniprojekt in einem temporaeren Verzeichnis."""
    root = tmp_path / "gravel_tracking"
    (root / "config").mkdir(parents=True)
    shutil.copy(PROJECT_ROOT / "config" / "conversion_factors.yaml", root / "config" / "conversion_factors.yaml")
    shutil.copytree(PROJECT_ROOT / "config" / "supplier_templates", root / "config" / "supplier_templates")

    shutil.copy(PROJECT_ROOT / "config" / "lv_mapping.yaml", root / "config" / "lv_mapping.yaml")
    config_text = (PROJECT_ROOT / "config" / "config.yaml").read_text(encoding="utf-8")
    config_text = config_text.replace(
        'lv_file: "data/lv/02_Leistungsmeldung_2026_CM_DE_v02_3.xlsx"',
        'lv_file: "data/lv/leistungsmeldung.xlsx"',
    )
    # Der Fixture hat nur das Hauptblatt.
    config_text = config_text.replace(
        '    - name: "Nachtrags_Leistung"\n      header_row: 12\n      lv_source: "Nachtrags-LV"\n', ""
    )
    (root / "config" / "config.yaml").write_text(config_text, encoding="utf-8")

    write_erp_fixture(root / "data" / "erp" / "test_receipts.xlsx")
    write_lv_fixture(root / "data" / "lv" / "leistungsmeldung.xlsx")
    (root / "data" / "gravel_deliveries" / "_unsorted").mkdir(parents=True)
    return root


@pytest.fixture(autouse=True)
def _clear_config_cache():
    from src.config import load_config, load_conversion_factors

    load_config.cache_clear()
    load_conversion_factors.cache_clear()
    yield
    load_config.cache_clear()
    load_conversion_factors.cache_clear()


def run_cli(root: Path, *args: str) -> int:
    from src.cli import main

    return main(["--config", str(root / "config" / "config.yaml"), *args])
