"""Formatvorgaben der Power BI Arbeitsmappe."""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from src.excel_out import verify_workbook
from src.tasks.model_build import (
    DIM_AREA,
    DIM_DATE,
    DIM_LOCATION,
    DIM_LV_POSITION,
    DIM_MATERIAL,
    DIM_MATERIAL_GROUP,
    DIM_SUPPLIER,
    FACT_DELIVERY,
    FACT_LOCATION_ALLOCATION,
    FACT_LV_BILLING,
)
from tests.conftest import run_cli

EXPECTED = {
    "fact_delivery": FACT_DELIVERY,
    "fact_lv_billing": FACT_LV_BILLING,
    "dim_area": DIM_AREA,
    "dim_date": DIM_DATE,
    "dim_material": DIM_MATERIAL,
    "dim_material_group": DIM_MATERIAL_GROUP,
    "dim_supplier": DIM_SUPPLIER,
    "dim_lv_position": DIM_LV_POSITION,
    "fact_location_allocation": FACT_LOCATION_ALLOCATION,
    "dim_location": DIM_LOCATION,
}


def _workbook(project: Path) -> Path:
    run_cli(project, "run", "--until-done")
    return project / "outputs" / "powerbi" / "gravel_model.xlsx"


def test_mappe_erfuellt_die_formatvorgaben(project: Path):
    problems = verify_workbook(_workbook(project), EXPECTED)
    assert problems == []


def test_jedes_blatt_traegt_genau_eine_tabelle_mit_blattnamen(project: Path):
    wb = load_workbook(_workbook(project))
    assert set(wb.sheetnames) == set(EXPECTED)
    for name in wb.sheetnames:
        ws = wb[name]
        tables = list(ws.tables.values())
        assert len(tables) == 1
        assert tables[0].displayName == name
        assert ws.cell(row=1, column=1).value == EXPECTED[name][0]


def test_datumswerte_sind_echte_datumswerte(project: Path):
    wb = load_workbook(_workbook(project))
    ws = wb["dim_date"]
    assert isinstance(ws.cell(row=2, column=1).value, (date, datetime))
    fact = wb["fact_delivery"]
    header = [c.value for c in fact[1]]
    col = header.index("delivery_date") + 1
    assert isinstance(fact.cell(row=2, column=col).value, (date, datetime))


def test_dim_date_ist_lueckenlos(project: Path):
    wb = load_workbook(_workbook(project))
    values = [row[0] for row in wb["dim_date"].iter_rows(min_row=2, max_col=1, values_only=True)]
    days = [v.date() if isinstance(v, datetime) else v for v in values]
    assert days == sorted(days)
    assert all((days[i + 1] - days[i]).days == 1 for i in range(len(days) - 1))


def test_keine_leerzeilen_in_der_faktentabelle(project: Path):
    wb = load_workbook(_workbook(project))
    ws = wb["fact_delivery"]
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        assert row[0] not in (None, "")


def test_spaltennamen_sind_englisch_ohne_umlaute_und_leerzeichen():
    for columns in EXPECTED.values():
        for column in columns:
            assert column == column.lower()
            assert " " not in column
            assert not set("äöüßÄÖÜ") & set(column)
